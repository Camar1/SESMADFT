from __future__ import annotations

import math
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .calculations import laminate_sequence, lttb_reduce, panel_abbreviation


LABELS = {
    "type": "Type",
    "date": "Date",
    "cf_type": "CF type",
    "laminate": "Laminate",
    "core_material": "Core material",
    "core_thickness_mm": "Core Thickness (mm)",
    "height_mm": "Height (mm)",
    "width_mm": "Width (mm)",
    "thickness_mm": "Thickness (mm)",
    "weight_g": "Weight (g)",
    "skin_thickness_mm": "Skin thickness (mm)",
}


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.\-eE+]", "", text)
    if not text:
        return None
    try:
        out = float(text)
        return out if math.isfinite(out) else None
    except ValueError:
        return None


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized(value: Any) -> str:
    return re.sub(r"\s+", " ", cell_to_string(value)).strip().lower()


def normalized_test_type(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", cell_to_string(value).lower())


def special_test_type_key(value: Any) -> str:
    text = normalized_test_type(value)
    if "harness" in text:
        return "harness"
    if "lapbelt" in text and ("parallel" in text or "paralel" in text):
        return "lap_belt_parallel"
    if "lapbelt" in text and ("perpendicular" in text or "perp" in text):
        return "lap_belt_perpendicular"
    return ""


def special_test_limit_kn(value: Any) -> float | None:
    key = special_test_type_key(value)
    if key == "harness":
        return 13.0
    if key in {"lap_belt_parallel", "lap_belt_perpendicular"}:
        return 19.5
    return None


def is_energy_absorbed_test(value: Any) -> bool:
    text = normalized_test_type(value)
    return (
        "3pb" in text
        or "3pointbend" in text
        or "threepointbend" in text
        or "sideimpact" in text
        or "sisv" in text
        or "sish" in text
    )


def energy_absorbed_value(specimen: dict[str, Any], computed: dict[str, Any], common: dict[str, Any]) -> float | None:
    candidates = (
        specimen.get("energyAbsorbedJ"),
        specimen.get("energy_absorbed_j"),
        specimen.get("energyAbsorbed"),
        common.get("energyAbsorbedJ"),
        common.get("energy_absorbed_j"),
        common.get("energyAbsorbed"),
        computed.get("energyAbsorbedJ"),
        computed.get("energy_absorbed_j"),
        computed.get("energyAbsorbed"),
    )
    for value in candidates:
        number = to_number(value)
        if number is not None:
            return number
    return None


def find_value_by_label(rows: list[list[Any]], label: str) -> str:
    target = normalized(label)
    for row in rows:
        for index, cell in enumerate(row):
            if normalized(cell) == target:
                for candidate in row[index + 1 :]:
                    text = cell_to_string(candidate)
                    if text:
                        return text
    return ""


def extract_metadata(rows: list[list[Any]]) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    for key, label in LABELS.items():
        value = find_value_by_label(rows, label)
        if key.endswith("_mm") or key == "weight_g":
            meta[key] = to_number(value)
        else:
            meta[key] = value

    meta["specimen_id"] = ""
    for row in rows[:6]:
        for cell in row:
            text = cell_to_string(cell)
            if re.match(r"^\d{4}-P\d{2}-\d{2}", text):
                meta["specimen_id"] = text
                return meta
    return meta


def normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", cell_to_string(value).lower())


def load_multiplier_from_header(value: Any) -> float:
    return 1000.0 if "kn" in normalized_header(value) else 1.0


def is_load_header(value: Any) -> bool:
    label = normalized_header(value)
    return label in {"load", "loadn", "force", "forcen", "y", "yn", "f", "fn", "fkn", "loadkn", "forcekn"}


def is_displacement_header(value: Any) -> bool:
    label = normalized_header(value)
    return label in {"displacement", "displacementmm", "x", "xmm", "disp", "dispmm", "s", "smm"}


def find_header_columns(rows: list[list[Any]]) -> tuple[int, int, int, float]:
    test_row = None
    for r_index, row in enumerate(rows):
        if any(normalized(cell) == "test data" for cell in row):
            test_row = r_index
            break

    candidates: list[int]
    if test_row is not None:
        candidates = [test_row, min(test_row + 1, len(rows) - 1)]
    else:
        candidates = list(range(min(len(rows), 140)))

    for r_index in candidates:
        row = rows[r_index]
        for c_index, cell in enumerate(row):
            if not is_load_header(cell):
                continue
            for disp_index in range(c_index + 1, min(len(row), c_index + 4)):
                if is_displacement_header(row[disp_index]):
                    return r_index, c_index, disp_index, load_multiplier_from_header(cell)
        load_col = None
        disp_col = None
        load_multiplier = 1.0
        for c_index, cell in enumerate(row):
            if is_load_header(cell):
                load_col = c_index
                load_multiplier = load_multiplier_from_header(cell)
            if is_displacement_header(cell):
                disp_col = c_index
        if load_col is not None and disp_col is not None:
            return r_index, load_col, disp_col, load_multiplier
    raise ValueError("LOAD / Displacement columns not found")


def extract_energy_target_mm(rows: list[list[Any]], header_row: int, load_col: int) -> float | None:
    for row in rows[max(0, header_row - 30) : header_row + 1]:
        for c_index, cell in enumerate(row):
            if "mmdeflection" not in normalized_header(cell):
                continue
            start = max(0, min(c_index, load_col + 8) - 8)
            end = min(len(row), c_index)
            for candidate in reversed(row[start:end]):
                number = to_number(candidate)
                if number is not None:
                    return number
    return None


def extract_load_displacement_info(rows: list[list[Any]]) -> dict[str, Any]:
    header_row, load_col, disp_col, load_multiplier = find_header_columns(rows)
    points: list[dict[str, float]] = []
    started = False
    for row in rows[header_row + 1 :]:
        load = row[load_col] if load_col < len(row) else None
        displacement = row[disp_col] if disp_col < len(row) else None
        if cell_to_string(load) == "" and cell_to_string(displacement) == "":
            if started:
                break
            continue
        y = to_number(load)
        x = to_number(displacement)
        if x is not None and y is not None:
            started = True
            points.append({"x_mm": x, "y_n": y * load_multiplier})
    if len(points) < 3:
        raise ValueError("Not enough load-displacement points found")
    return {
        "points": points,
        "header_row": header_row,
        "load_col": load_col,
        "disp_col": disp_col,
        "load_multiplier": load_multiplier,
        "energy_target_mm": extract_energy_target_mm(rows, header_row, load_col),
    }


def extract_load_displacement(rows: list[list[Any]]) -> list[dict[str, float]]:
    info = extract_load_displacement_info(rows)
    return info["points"]
    return points


def workbook_rows(workbook: openpyxl.Workbook, sheet_name: str | None = None) -> tuple[str, list[list[Any]]]:
    ws = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    return ws.title, rows


def parse_workbook_bytes(contents: bytes, sheet_name: str | None = None, include_metadata: bool = True) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    errors = []
    sheet_names = [sheet_name] if sheet_name else workbook.sheetnames
    for name in sheet_names:
        try:
            actual, rows = workbook_rows(workbook, name)
            load_info = extract_load_displacement_info(rows)
            metadata = extract_metadata(rows) if include_metadata else {}
            if load_info.get("energy_target_mm") is not None:
                metadata["energy_target_mm"] = load_info["energy_target_mm"]
            return {
                "sheet_name": actual,
                "metadata": metadata,
                "points": load_info["points"],
            }
        except Exception as exc:  # Try the next sheet for generic workbooks.
            errors.append(f"{name}: {exc}")
    raise ValueError("; ".join(errors))


def fmt_number(value: Any, digits: int = 4) -> Any:
    if isinstance(value, (int, float)) and math.isfinite(value):
        return round(float(value), digits)
    return value


def autosize(ws) -> None:
    for column_index, column in enumerate(ws.columns, start=1):
        letter = get_column_letter(column_index)
        width = min(max(len(str(cell.value or "")) + 2 for cell in column), 48)
        ws.column_dimensions[letter].width = max(width, 10)


def write_key_values(ws, start_row: int, title: str, rows: list[tuple[str, Any]]) -> int:
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="1F4E5F")
    ws.cell(start_row, 2, "")
    ws.cell(start_row, 2).fill = PatternFill("solid", fgColor="1F4E5F")
    row_index = start_row + 1
    for label, value in rows:
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, fmt_number(value))
        row_index += 1
    return row_index + 1


def export_specimen_workbook(specimen: dict[str, Any], output_path: Path) -> Path:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    data = wb.create_sheet("Processed Data")
    panels = wb.create_sheet("Panel Results")

    theoretical = specimen.get("theoretical", {}) or {}
    computed = specimen.get("computed", {}) or {}
    common = computed.get("common", {}) or {}
    limit_kn = special_test_limit_kn(specimen.get("test_type"))
    is_special = limit_kn is not None
    energy_absorbed = energy_absorbed_value(specimen, computed, common)
    energy_applicable = is_energy_absorbed_test(specimen.get("test_type")) or energy_absorbed is not None
    selected_points = [] if is_special else specimen.get("selected_points", []) or computed.get("selected_points", []) or []
    reduced_points = (
        specimen.get("reduced_points")
        or computed.get("reduced_points")
        or lttb_reduce(specimen.get("raw_points", []) or [], threshold=300)
    )

    row = 1
    specimen_rows = [
        ("ID", specimen.get("id")),
        ("Test type", specimen.get("test_type")),
        ("Laminate sequence", laminate_sequence(specimen.get("laminate")) if laminate_sequence(specimen.get("laminate")) != "Incomplete laminate" else theoretical.get("laminate_sequence")),
        ("Source file", specimen.get("source_filename")),
        ("Comments", specimen.get("comments")),
        ("Real weight [g]", specimen.get("real_weight_g")),
        ("Real thickness [mm]", specimen.get("real_thickness_mm")),
        ("Theoretical weight [g]", theoretical.get("total_weight_g")),
        ("Theoretical thickness [mm]", theoretical.get("total_thickness_mm")),
        ("0 deg fibers [%]", (theoretical.get("zero_percent") or 0) * 100 if theoretical.get("zero_percent") is not None else None),
        ("Warping [%]", (theoretical.get("warping_percent") or 0) * 100 if theoretical.get("warping_percent") is not None else None),
        ("Top skin thickness [mm]", computed.get("top_skin_thickness_mm") or specimen.get("top_skin_thickness_mm") or theoretical.get("top_skin_thickness_mm")),
        ("Bottom skin thickness [mm]", computed.get("bottom_skin_thickness_mm") or specimen.get("bottom_skin_thickness_mm") or theoretical.get("bottom_skin_thickness_mm")),
        ("Equivalent skin thickness [mm]", computed.get("skin_thickness_each_side_mm")),
        ("Core thickness [mm]", computed.get("core_thickness_mm") or specimen.get("core_thickness_mm")),
        ("Energy Absorbed [J]", energy_absorbed if energy_applicable else None),
    ]
    if is_special:
        specimen_rows.append(("Limit [kN]", limit_kn))
    row = write_key_values(
        summary,
        row,
        "Specimen",
        specimen_rows,
    )

    if is_special:
        row = write_key_values(
            summary,
            row,
            "Special Test",
            [
                ("Limit [kN]", limit_kn),
                ("Fmax [N]", common.get("max_force_n")),
            ],
        )
    elif computed.get("mode") == "Shear":
        row = write_key_values(
            summary,
            row,
            "Shear Points",
            [
                ("Higher Peak x [mm]", common.get("highest_peak_x_mm")),
                ("Higher Peak y [N]", common.get("highest_peak_n")),
                ("Lower Peak x [mm]", common.get("lowest_peak_x_mm")),
                ("Lower Peak y [N]", common.get("lowest_peak_n")),
            ],
        )
    else:
        region = common.get("linear_region", {}) if isinstance(common, dict) else {}
        row = write_key_values(
            summary,
            row,
            "3PB Points",
            [
                ("x1 [mm]", region.get("x1_mm")),
                ("x2 [mm]", region.get("x2_mm")),
                ("y1 [N]", region.get("y1_n")),
                ("y2 [N]", region.get("y2_n")),
                ("Slope [N/mm]", region.get("slope_n_per_mm")),
                ("Fmax [N]", common.get("max_force_n")),
                ("Energy Absorbed [J]", energy_absorbed),
            ],
        )

    summary.cell(row, 1, "Reduced Curve (300 points)")
    summary.cell(row, 1).font = Font(bold=True, color="FFFFFF")
    summary.cell(row, 1).fill = PatternFill("solid", fgColor="1F4E5F")
    row += 1
    summary.cell(row, 1, "Displacement [mm]")
    summary.cell(row, 2, "Load [kN]")
    for col in (1, 2):
        summary.cell(row, col).font = Font(bold=True)
        summary.cell(row, col).fill = PatternFill("solid", fgColor="E7EEF2")
    row += 1
    for point in reduced_points:
        y_kn = point.get("y_n") / 1000.0 if isinstance(point.get("y_n"), (int, float)) else None
        summary.cell(row, 1, point.get("x_mm"))
        summary.cell(row, 2, y_kn)
        row += 1

    if is_special:
        panels.append(["Special test", "Panel safety coefficients are not applicable"])
    else:
        panels.append(
            [
                "Panel",
                "EI [GPa*m^4]",
                "CS EI",
                "Yield [N]",
                "CS Yield",
                "UTS [N]",
                "CS UTS",
                "Max Load Midspan [N]",
                "CS Max Load",
                "Max Deflection [m]",
                "CS Deflection",
                "Energy [J]",
                "CS Energy",
                "Shear Stress [MPa]",
                "Perimeter Shear [N]",
                "CS Perimeter Shear",
            ]
        )
        for panel_code, values in (computed.get("panel_results") or {}).items():
            panels.append(
                [
                    panel_abbreviation(panel_code),
                    values.get("ei_gpa_m4"),
                    values.get("cs_ei"),
                    values.get("yield_probe_n"),
                    values.get("cs_yield"),
                    values.get("uts_probe_n"),
                    values.get("cs_uts"),
                    values.get("max_load_midspan_probe_n"),
                    values.get("cs_max_load_midspan"),
                    values.get("max_deflection_probe_m"),
                    values.get("cs_max_deflection"),
                    values.get("energy_absorbed_probe_j"),
                    values.get("cs_energy"),
                    values.get("shear_stress_mpa"),
                    values.get("perimeter_shear_probe_n"),
                    values.get("cs_perimeter_shear"),
                ]
            )

    if is_special:
        data.append(["x [mm]", "Load [kN]", None, "Limit x [mm]", "Limit load [kN]"])
    else:
        data.append(["x [mm]", "Load [kN]", None, "Selected role", "Selected x [mm]", "Selected y [kN]"])
    for point in reduced_points:
        y_kn = point.get("y_n") / 1000.0 if isinstance(point.get("y_n"), (int, float)) else None
        data.append([point.get("x_mm"), y_kn])
    limit_rows_written = False
    if is_special and limit_kn is not None:
        x_values_for_limit = [
            point.get("x_mm")
            for point in reduced_points
            if isinstance(point.get("x_mm"), (int, float))
        ]
        if x_values_for_limit:
            data.cell(2, 4, min(x_values_for_limit))
            data.cell(2, 5, limit_kn)
            data.cell(3, 4, max(x_values_for_limit))
            data.cell(3, 5, limit_kn)
            limit_rows_written = True
    else:
        selected_start = 2
        for offset, point in enumerate(selected_points):
            r = selected_start + offset
            data.cell(r, 4, point.get("role"))
            data.cell(r, 5, point.get("x_mm"))
            y_kn = point.get("y_n") / 1000.0 if isinstance(point.get("y_n"), (int, float)) else None
            data.cell(r, 6, y_kn)

    if data.max_row > 2:
        chart = ScatterChart()
        chart.title = "Load - Displacement"
        chart.x_axis.title = "Displacement [mm]"
        chart.y_axis.title = "Load [kN]"
        chart.x_axis.majorUnit = 1
        chart.y_axis.majorUnit = 5
        chart.height = 10
        chart.width = 18
        x_values = Reference(data, min_col=1, min_row=2, max_row=data.max_row)
        y_values = Reference(data, min_col=2, min_row=2, max_row=data.max_row)
        series = Series(y_values, x_values, title="Curve")
        series.graphicalProperties.line.solidFill = "111111"
        series.marker.symbol = "none"
        chart.series.append(series)
        if selected_points:
            sx = Reference(data, min_col=5, min_row=2, max_row=1 + len(selected_points))
            sy = Reference(data, min_col=6, min_row=2, max_row=1 + len(selected_points))
            selected_series = Series(sy, sx, title="Selected points")
            selected_series.graphicalProperties.line.noFill = True
            selected_series.marker.symbol = "circle"
            selected_series.marker.size = 8
            selected_series.marker.graphicalProperties.solidFill = "F59E0B"
            selected_series.marker.graphicalProperties.line.solidFill = "9A3412"
            chart.series.append(selected_series)
        if limit_rows_written:
            lx = Reference(data, min_col=4, min_row=2, max_row=3)
            ly = Reference(data, min_col=5, min_row=2, max_row=3)
            limit_series = Series(ly, lx, title=f"Limit: {fmt_number(limit_kn)} kN")
            limit_series.graphicalProperties.line.solidFill = "FF4F00"
            limit_series.marker.symbol = "none"
            chart.series.append(limit_series)
        summary.add_chart(chart, "D2")

    for ws in [summary, panels, data]:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E7EEF2")
            cell.alignment = Alignment(vertical="center")
        autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
